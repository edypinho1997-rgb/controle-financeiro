import json
import os
from functools import partial

from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.scrollview import ScrollView
from kivy.uix.widget import Widget
from kivy.core.window import Window
from kivy.graphics import Color, Rectangle
from kivy.properties import NumericProperty

from kivymd.uix.label import MDLabel
from kivymd.uix.textfield import MDTextField
from kivymd.uix.button import MDRaisedButton



# 🎨 CORES
CARD = (0.10, 0.14, 0.18, 1)
VERDE = (0.2, 0.8, 0.4, 1)
VERMELHO = (0.9, 0.2, 0.2, 1)
AZUL = (0.2, 0.6, 0.9, 1)
BRANCO = (1, 1, 1, 1)
ROSA = (0.95, 0.62, 0.78, 1)
LARGURA_MOBILE = 500


def formatar_real(valor):
    return f"R${valor:.2f}"


def ler_valor(texto):
    texto = texto.strip().replace("R$", "").replace(" ", "")
    if not texto:
        return 0
    texto = texto.replace(".", "").replace(",", ".")
    return float(texto)


def caminho_app(nome_arquivo):
    app = MDApp.get_running_app()
    base_dir = app.user_data_dir if app else os.getcwd()
    os.makedirs(base_dir, exist_ok=True)
    return os.path.join(base_dir, nome_arquivo)


def preparar_label(label):
    label.bind(width=lambda instance, value: setattr(instance, "text_size", (value, None)))
    return label


class Card(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        with self.canvas.before:
            Color(*CARD)
            self.rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update, size=self.update)

    def update(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size


class Grafico(Widget):
    entrada = NumericProperty(0)
    saida = NumericProperty(0)
    investimento = NumericProperty(0)

    def on_size(self, *args): self.desenhar()
    def on_entrada(self, *args): self.desenhar()
    def on_saida(self, *args): self.desenhar()
    def on_investimento(self, *args): self.desenhar()

    def desenhar(self):
        if self.canvas is None:
            return

        self.canvas.clear()
        total = self.entrada if self.entrada > 0 else 1

        ps = self.saida / total
        pi = self.investimento / total
        pe = 1 - ps - pi

        if pe < 0:
            pe = 0

        with self.canvas:
            Color(*VERDE)
            Rectangle(pos=self.pos, size=(self.width * pe, self.height * 0.5))

            Color(*VERMELHO)
            Rectangle(pos=(self.x + self.width * pe, self.y),
                      size=(self.width * ps, self.height * 0.5))

            Color(*AZUL)
            Rectangle(pos=(self.x + self.width * (pe + ps), self.y),
                      size=(self.width * pi, self.height * 0.5))


class TelaPrincipal(Screen):
    entrada = NumericProperty(0)
    saida = NumericProperty(0)
    investimento = NumericProperty(0)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.mobile = Window.width <= LARGURA_MOBILE
        self.arquivo = caminho_app("dados.json")
        self.arquivo_excel = caminho_app("financeiro.xlsx")
        self.dados = []
        self.maria_cecilia = []
        self.carregar_dados()

        scroll = ScrollView(do_scroll_x=False)
        layout = GridLayout(
            cols=1,
            padding=10 if self.mobile else 15,
            spacing=10 if self.mobile else 15,
            size_hint_y=None,
        )
        layout.bind(minimum_height=layout.setter("height"))

        # RESUMO
        q1 = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        q1.bind(minimum_height=q1.setter("height"))
        self.label_total = preparar_label(MDLabel(
            text="", halign="center", theme_text_color="Custom", text_color=BRANCO, adaptive_height=True
        ))
        q1.add_widget(self.label_total)

        btn_hist = MDRaisedButton(text="HISTÓRICO")
        btn_hist.bind(on_release=self.ir_historico)
        q1.add_widget(btn_hist)

        btn_excel = MDRaisedButton(text="EXPORTAR EXCEL")
        btn_excel.bind(on_release=self.exportar_excel)
        q1.add_widget(btn_excel)

        btn_maria = MDRaisedButton(text="MARIA CECILIA")
        btn_maria.md_bg_color = ROSA
        btn_maria.bind(on_release=self.ir_maria_cecilia)
        q1.add_widget(btn_maria)

        # GRAFICO
        q2 = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        q2.bind(minimum_height=q2.setter("height"))
        self.label_info = preparar_label(MDLabel(
            text="", halign="center", theme_text_color="Custom", text_color=BRANCO, adaptive_height=True
        ))
        q2.add_widget(self.label_info)

        self.grafico = Grafico(size_hint_y=None, height=90 if self.mobile else 120)
        q2.add_widget(self.grafico)

        # ENTRADA
        q3 = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        q3.bind(minimum_height=q3.setter("height"))
        q3.add_widget(MDLabel(text="ENTRADA", theme_text_color="Custom", text_color=VERDE))

        self.nome_entrada = MDTextField(hint_text="Nome")
        self.data_entrada = MDTextField(hint_text="Data")
        self.valor_entrada = MDTextField(hint_text="Valor")

        btn_entrada = MDRaisedButton(text="Adicionar Entrada")
        btn_entrada.md_bg_color = VERDE
        btn_entrada.bind(on_release=self.add_entrada)

        q3.add_widget(self.nome_entrada)
        q3.add_widget(self.data_entrada)
        q3.add_widget(self.valor_entrada)
        q3.add_widget(btn_entrada)

        # SAIDA
        q4 = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        q4.bind(minimum_height=q4.setter("height"))
        q4.add_widget(MDLabel(text="SAÍDA", theme_text_color="Custom", text_color=VERMELHO))

        self.nome_saida = MDTextField(hint_text="Nome")
        self.data_saida = MDTextField(hint_text="Data")
        self.valor_saida = MDTextField(hint_text="Valor")

        btn_saida = MDRaisedButton(text="Adicionar Saída")
        btn_saida.md_bg_color = VERMELHO
        btn_saida.bind(on_release=self.add_saida)

        q4.add_widget(self.nome_saida)
        q4.add_widget(self.data_saida)
        q4.add_widget(self.valor_saida)
        q4.add_widget(btn_saida)

        # INVESTIMENTO
        q5 = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        q5.bind(minimum_height=q5.setter("height"))
        q5.add_widget(MDLabel(text="INVESTIMENTOS", theme_text_color="Custom", text_color=AZUL))

        self.nome_investimento = MDTextField(hint_text="Nome")
        self.data_investimento = MDTextField(hint_text="Data")
        self.valor_investimento = MDTextField(hint_text="Valor")

        btn_investimento = MDRaisedButton(text="Adicionar Investimento")
        btn_investimento.md_bg_color = AZUL
        btn_investimento.bind(on_release=self.add_investimento)

        q5.add_widget(self.nome_investimento)
        q5.add_widget(self.data_investimento)
        q5.add_widget(self.valor_investimento)
        q5.add_widget(btn_investimento)

        layout.add_widget(q1)
        layout.add_widget(q2)
        layout.add_widget(q3)
        layout.add_widget(q4)
        layout.add_widget(q5)

        scroll.add_widget(layout)
        self.add_widget(scroll)
        self.atualizar()

    # EXPORTAR EXCEL COM TOTAL
    def exportar_excel(self, instance):
        from openpyxl import Workbook

        wb = Workbook()

        ws_entrada = wb.active
        ws_entrada.title = "Entradas"
        ws_entrada.append(["Nome", "Data", "Valor"])

        ws_saida = wb.create_sheet(title="Saidas")
        ws_saida.append(["Nome", "Data", "Valor"])

        ws_investimento = wb.create_sheet(title="Investimentos")
        ws_investimento.append(["Nome", "Data", "Valor"])

        total_entrada = 0
        total_saida = 0
        total_investimento = 0

        for item in self.dados:
            if item["tipo"] == "entrada":
                ws_entrada.append([item["nome"], item["data"], item["valor"]])
                total_entrada += item["valor"]
            elif item["tipo"] == "saida":
                ws_saida.append([item["nome"], item["data"], item["valor"]])
                total_saida += item["valor"]
            else:
                ws_investimento.append([item["nome"], item["data"], item["valor"]])
                total_investimento += item["valor"]

        ws_entrada.append([])
        ws_entrada.append(["TOTAL", "", total_entrada])

        ws_saida.append([])
        ws_saida.append(["TOTAL", "", total_saida])

        ws_investimento.append([])
        ws_investimento.append(["TOTAL", "", total_investimento])

        wb.save(self.arquivo_excel)
        print("Excel criado com total!")

    def salvar_dados(self):
        with open(self.arquivo, "w", encoding="utf-8") as f:
            json.dump({
                "entrada": self.entrada,
                "saida": self.saida,
                "investimento": self.investimento,
                "dados": self.dados,
                "maria_cecilia": self.maria_cecilia,
            }, f)

    def carregar_dados(self):
        if os.path.exists(self.arquivo):
            try:
                with open(self.arquivo, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.entrada = data.get("entrada", 0)
                    self.saida = data.get("saida", 0)
                    self.investimento = data.get("investimento", 0)
                    self.dados = data.get("dados", [])
                    self.maria_cecilia = data.get("maria_cecilia", [])
            except (json.JSONDecodeError, OSError):
                self.entrada = 0
                self.saida = 0
                self.investimento = 0
                self.dados = []
                self.maria_cecilia = []

    def atualizar(self):
        total = self.entrada if self.entrada > 0 else 1
        ps = (self.saida / total) * 100
        pi = (self.investimento / total) * 100
        pe = max(0, 100 - ps - pi)
        saldo = self.entrada - self.saida - self.investimento

        self.label_total.text = f"Saldo: {formatar_real(saldo)}"
        if self.mobile:
            self.label_info.text = (
                f"Entrada: {formatar_real(self.entrada)} ({pe:.0f}%)\n"
                f"Saida: {formatar_real(self.saida)} ({ps:.0f}%)\n"
                f"Invest.: {formatar_real(self.investimento)} ({pi:.0f}%)\n"
                f"Resultado: {formatar_real(saldo)}"
            )
        else:
            self.label_info.text = (
                f"Entrada: {formatar_real(self.entrada)} | "
                f"Saida: {formatar_real(self.saida)} | "
                f"Invest.: {formatar_real(self.investimento)} | "
                f"Resultado: {formatar_real(saldo)}"
            )

        self.grafico.entrada = self.entrada
        self.grafico.saida = self.saida
        self.grafico.investimento = self.investimento

    def add_entrada(self, instance):
        if self.valor_entrada.text:
            valor = float(self.valor_entrada.text)

            self.entrada += valor
            self.dados.append({
                "tipo": "entrada",
                "nome": self.nome_entrada.text,
                "data": self.data_entrada.text,
                "valor": valor
            })

            self.salvar_dados()
            self.nome_entrada.text = ""
            self.data_entrada.text = ""
            self.valor_entrada.text = ""
            self.atualizar()

    def add_saida(self, instance):
        if self.valor_saida.text:
            valor = float(self.valor_saida.text)

            self.saida += valor
            self.dados.append({
                "tipo": "saida",
                "nome": self.nome_saida.text,
                "data": self.data_saida.text,
                "valor": valor
            })

            self.salvar_dados()
            self.nome_saida.text = ""
            self.data_saida.text = ""
            self.valor_saida.text = ""
            self.atualizar()

    def add_investimento(self, instance):
        if self.valor_investimento.text:
            valor = float(self.valor_investimento.text)

            self.investimento += valor
            self.dados.append({
                "tipo": "investimento",
                "nome": self.nome_investimento.text,
                "data": self.data_investimento.text,
                "valor": valor
            })

            self.salvar_dados()
            self.nome_investimento.text = ""
            self.data_investimento.text = ""
            self.valor_investimento.text = ""
            self.atualizar()

    def ir_historico(self, instance):
        tela = self.manager.get_screen("historico")
        tela.mostrar(self.dados, self)
        self.manager.current = "historico"

    def ir_maria_cecilia(self, instance):
        tela = self.manager.get_screen("maria_cecilia")
        tela.mostrar(self)
        self.manager.current = "maria_cecilia"


class TelaHistorico(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.mobile = Window.width <= LARGURA_MOBILE
        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        scroll_principal = ScrollView(do_scroll_x=False)
        colunas = GridLayout(cols=1, spacing=12, size_hint_y=None)
        colunas.bind(minimum_height=colunas.setter("height"))

        self.col_entrada = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            padding=8 if self.mobile else 10,
            spacing=8,
        )
        self.col_entrada.bind(minimum_height=self.col_entrada.setter("height"))

        self.col_saida = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            padding=8 if self.mobile else 10,
            spacing=8,
        )
        self.col_saida.bind(minimum_height=self.col_saida.setter("height"))

        self.col_investimento = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            padding=8 if self.mobile else 10,
            spacing=8,
        )
        self.col_investimento.bind(minimum_height=self.col_investimento.setter("height"))

        card_entrada = Card(orientation="vertical", size_hint_y=None)
        card_entrada.bind(minimum_height=card_entrada.setter("height"))
        card_entrada.add_widget(self.col_entrada)

        card_saida = Card(orientation="vertical", size_hint_y=None)
        card_saida.bind(minimum_height=card_saida.setter("height"))
        card_saida.add_widget(self.col_saida)

        card_investimento = Card(orientation="vertical", size_hint_y=None)
        card_investimento.bind(minimum_height=card_investimento.setter("height"))
        card_investimento.add_widget(self.col_investimento)

        colunas.add_widget(card_entrada)
        colunas.add_widget(card_saida)
        colunas.add_widget(card_investimento)
        scroll_principal.add_widget(colunas)

        btn_voltar = MDRaisedButton(text="VOLTAR")
        btn_voltar.bind(on_release=self.voltar)

        layout.add_widget(scroll_principal)
        layout.add_widget(btn_voltar)

        self.add_widget(layout)

    def mostrar(self, dados, tela_principal):
        self.col_entrada.clear_widgets()
        self.col_saida.clear_widgets()
        self.col_investimento.clear_widgets()

        self.tela_principal = tela_principal

        self.col_entrada.add_widget(
            MDLabel(text="Entradas", theme_text_color="Custom", text_color=VERDE, size_hint_y=None, height=40)
        )
        self.col_saida.add_widget(
            MDLabel(text="Saidas", theme_text_color="Custom", text_color=VERMELHO, size_hint_y=None, height=40)
        )
        self.col_investimento.add_widget(
            MDLabel(text="Investimentos", theme_text_color="Custom", text_color=AZUL, size_hint_y=None, height=40)
        )

        for i, item in enumerate(dados):
            linha = BoxLayout(
                orientation="vertical" if self.mobile else "horizontal",
                size_hint_y=None,
                spacing=8,
            )
            linha.bind(minimum_height=linha.setter("height"))

            texto = f"{item['nome']} | {item['data']} | {formatar_real(item['valor'])}"
            if item["tipo"] == "entrada":
                cor = VERDE
            elif item["tipo"] == "saida":
                cor = VERMELHO
            else:
                cor = AZUL

            label = MDLabel(
                text=texto,
                theme_text_color="Custom",
                text_color=cor,
                adaptive_height=True,
            )
            preparar_label(label)

            btn = MDRaisedButton(
                text="X",
                size_hint_x=None if not self.mobile else 1,
                width=60,
            )
            btn.bind(on_release=partial(self.apagar, i))

            linha.add_widget(label)
            linha.add_widget(btn)

            if item["tipo"] == "entrada":
                self.col_entrada.add_widget(linha)
            elif item["tipo"] == "saida":
                self.col_saida.add_widget(linha)
            else:
                self.col_investimento.add_widget(linha)

    def apagar(self, index, *args):
        item = self.tela_principal.dados[index]

        if item["tipo"] == "entrada":
            self.tela_principal.entrada -= item["valor"]
        elif item["tipo"] == "saida":
            self.tela_principal.saida -= item["valor"]
        else:
            self.tela_principal.investimento -= item["valor"]

        del self.tela_principal.dados[index]

        self.tela_principal.salvar_dados()
        self.tela_principal.atualizar()

        self.mostrar(self.tela_principal.dados, self.tela_principal)

    def voltar(self, instance):
        self.manager.current = "principal"


class TelaMariaCecilia(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.mobile = Window.width <= LARGURA_MOBILE

        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        scroll_principal = ScrollView(do_scroll_x=False)
        conteudo = GridLayout(cols=1, spacing=12, size_hint_y=None)
        conteudo.bind(minimum_height=conteudo.setter("height"))

        card_form = Card(orientation="vertical", padding=15, spacing=12, size_hint_y=None)
        card_form.bind(minimum_height=card_form.setter("height"))
        card_form.add_widget(
            MDLabel(text="MARIA CECILIA", theme_text_color="Custom", text_color=ROSA, adaptive_height=True)
        )

        self.nome_compra = MDTextField(hint_text="Compra")

        btn_add = MDRaisedButton(text="Adicionar Compra")
        btn_add.md_bg_color = ROSA
        btn_add.bind(on_release=self.add_compra)

        self.label_total_compras = preparar_label(MDLabel(
            text="",
            halign="center",
            theme_text_color="Custom",
            text_color=BRANCO,
            adaptive_height=True,
        ))

        card_form.add_widget(self.label_total_compras)
        card_form.add_widget(self.nome_compra)
        card_form.add_widget(btn_add)

        card_lista = Card(orientation="vertical", padding=15, spacing=10, size_hint_y=None)
        card_lista.bind(minimum_height=card_lista.setter("height"))
        card_lista.add_widget(
            MDLabel(text="Compras Planejadas", theme_text_color="Custom", text_color=ROSA, adaptive_height=True)
        )

        self.lista_compras = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            spacing=8,
        )
        self.lista_compras.bind(minimum_height=self.lista_compras.setter("height"))
        card_lista.add_widget(self.lista_compras)

        conteudo.add_widget(card_form)
        conteudo.add_widget(card_lista)

        scroll_principal.add_widget(conteudo)

        btn_voltar = MDRaisedButton(text="VOLTAR")
        btn_voltar.bind(on_release=self.voltar)

        layout.add_widget(scroll_principal)
        layout.add_widget(btn_voltar)
        self.add_widget(layout)

    def mostrar(self, tela_principal):
        self.tela_principal = tela_principal
        self.atualizar_lista()

    def atualizar_lista(self):
        self.lista_compras.clear_widgets()
        compras = getattr(self.tela_principal, "maria_cecilia", [])
        total = sum(item["valor"] for item in compras)
        self.label_total_compras.text = f"Total planejado: {formatar_real(total)}"

        if not compras:
            self.lista_compras.add_widget(
                MDLabel(
                    text="Nenhuma compra cadastrada ainda.",
                    theme_text_color="Custom",
                    text_color=BRANCO,
                    adaptive_height=True,
                )
            )
            return

        for i, item in enumerate(compras):
            linha = BoxLayout(
                orientation="vertical" if self.mobile else "horizontal",
                size_hint_y=None,
                spacing=8,
                padding=(0, 4),
            )
            linha.bind(minimum_height=linha.setter("height"))

            label = preparar_label(MDLabel(
                text=item["nome"],
                theme_text_color="Custom",
                text_color=ROSA,
                adaptive_height=True,
                size_hint_x=1 if self.mobile else 0.55,
                valign="middle",
            ))

            checkbox = CheckBox(
                active=item.get("comprado", False),
                size_hint_x=None,
                width=40,
            )
            checkbox.bind(active=partial(self.marcar_compra, i))

            campo_valor = MDTextField(
                text="" if item.get("valor", 0) == 0 else f"{item['valor']:.2f}",
                hint_text="VALOR",
                size_hint_x=1 if self.mobile else 0.25,
            )
            campo_valor.bind(text=partial(self.atualizar_valor_compra, i))

            btn_remover = MDRaisedButton(
                text="X",
                size_hint_x=None if not self.mobile else 1,
                width=50 if not self.mobile else 0,
            )
            btn_remover.bind(on_release=partial(self.apagar_compra, i))

            if self.mobile:
                linha_controles = BoxLayout(
                    orientation="horizontal",
                    size_hint_y=None,
                    height=56,
                    spacing=8,
                )
                linha_controles.add_widget(checkbox)
                linha_controles.add_widget(campo_valor)
                linha_controles.add_widget(btn_remover)
                linha.add_widget(label)
                linha.add_widget(linha_controles)
            else:
                linha.add_widget(label)
                linha.add_widget(checkbox)
                linha.add_widget(campo_valor)
                linha.add_widget(btn_remover)

            self.lista_compras.add_widget(linha)

    def add_compra(self, instance):
        if self.nome_compra.text:
            self.tela_principal.maria_cecilia.append({
                "nome": self.nome_compra.text,
                "valor": 0,
                "comprado": False,
            })
            self.tela_principal.salvar_dados()
            self.nome_compra.text = ""
            self.atualizar_lista()

    def marcar_compra(self, index, checkbox, ativo):
        self.tela_principal.maria_cecilia[index]["comprado"] = ativo
        self.tela_principal.salvar_dados()

    def apagar_compra(self, index, *args):
        del self.tela_principal.maria_cecilia[index]
        self.tela_principal.salvar_dados()
        self.atualizar_lista()

    def atualizar_valor_compra(self, index, campo, texto):
        try:
            valor = ler_valor(texto)
        except ValueError:
            return

        self.tela_principal.maria_cecilia[index]["valor"] = valor
        self.tela_principal.salvar_dados()
        total = sum(item["valor"] for item in self.tela_principal.maria_cecilia)
        self.label_total_compras.text = f"Total planejado: {formatar_real(total)}"

    def voltar(self, instance):
        self.manager.current = "principal"


class MeuApp(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"

        sm = ScreenManager()
        sm.add_widget(TelaPrincipal(name="principal"))
        sm.add_widget(TelaHistorico(name="historico"))
        sm.add_widget(TelaMariaCecilia(name="maria_cecilia"))

        Window.bind(on_keyboard=self.on_back_button)
        return sm

    def on_back_button(self, window, key, scancode, codepoint, modifiers):
        if key != 27:
            return False

        if self.root.current == "historico":
            self.root.current = "principal"
            return True

        if self.root.current == "maria_cecilia":
            self.root.current = "principal"
            return True

        return True


MeuApp().run()
